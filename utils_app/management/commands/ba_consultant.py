import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.core.management import BaseCommand

from project.models import Project


class Command(BaseCommand):
    @staticmethod
    def create_csv(data):
        # Convert dictionary to tabular data
        rows = []
        for consultant, records in data.items():
            marketing = consultant.marketing.filter().order_by("-id").first()
            preferred_location = marketing.preferred_location if marketing else "NA"
            for record in records:
                rows.append([
                    consultant.id, consultant.name, consultant.email, consultant.phone_no,
                    "No" if consultant.status.lower() == 'terminated' else "Yes",
                    consultant.skills, consultant.current_city, preferred_location,
                    record["po_id"], record["client"], record["vendor_company"],
                    record["job_position"], record["job_title"],
                    record["work_type"], record["po_status"], record["log1_url"]
                ])

        # Create a DataFrame
        df = pd.DataFrame(rows, columns=[
            "ID", "Consultant Name", "Consultant Email", "Consultant Contact Details", "Consultant Is Active",
            "Consultant Skills", "Consultant Location", "Consultant Preferred Location",
            "PO ID", "Client", "Vendor Company", "Job Position", "Job Title", "Work Type", "PO Status", "Log1 URL"
        ])

        # Create an Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grouped Data"

        # Add headers
        headers = [
            "ID", "Consultant Name", "Consultant Email", "Consultant Contact Details", "Consultant Is Active",
            "Consultant Skills", "Consultant Location", "Consultant Preferred Location",
            "PO ID", "Client", "Vendor Company", "Job Position", "Job Title", "Work Type", "PO Status", "Log1 URL"
        ]
        ws.append(headers)

        # Add data with merged cells based on "ID" column
        merge_up_to_column = headers.index("Consultant Preferred Location") + 1
        start_row = 2
        current_id = None

        for idx, row in df.iterrows():
            current_row = idx + 2  # Account for header row
            row_id = row["ID"]

            # Check if the ID changes
            if row_id != current_id:
                # Merge the previous group
                if current_id is not None:
                    for col_index in range(merge_up_to_column):
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_index + 1,
                            end_row=current_row - 1,
                            end_column=col_index + 1,
                        )
                current_id = row_id
                start_row = current_row

            # Add row to worksheet
            ws.append(row.tolist())

            # Add hyperlink for Log1 URL
            log1_url_formula = f'=HYPERLINK("{row["Log1 URL"]}", "Log1 Redirect URL")'
            ws.cell(row=current_row, column=len(headers)).value = log1_url_formula

        # Merge the last group of cells
        for col_index in range(merge_up_to_column):
            ws.merge_cells(
                start_row=start_row,
                start_column=col_index + 1,
                end_row=len(df) + 1,
                end_column=col_index + 1,
            )

        # Align cells
        for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center")

        # Save the Excel file
        wb.save("BA_Consultant.xlsx")

    def handle(self, *args, **options):
        try:
            data = {}
            position_ids = [68, 69, 70, 71, 72, 73, 74]
            project_qs = Project.objects.filter(submission__lead__position_id__in=position_ids).order_by("consultant")
            for obj in project_qs:
                consultant_obj = obj.submission.consultant
                if consultant_obj not in data:
                    data[consultant_obj] = []

                data[consultant_obj].append({
                    "po_id": obj.id, "client": obj.submission.client, "vendor_company": obj.submission.vendor.name,
                    "job_position": obj.submission.lead.position.display_name, "job_title": obj.submission.lead.job_title,
                    "log1_url": f"https://log1.com/#/details/{obj.submission_id}/project", "po_status": obj.status,
                    "work_type": obj.submission.get_work_type_display()
                })
            self.create_csv(data)
        except Exception as error:
            self.stderr.write(f"Error: {error}")
